"""Excel reader for repos catalog."""

from pathlib import Path

from openpyxl import load_workbook


class OpenpyxlExcelReader:
    def read_rows(self, file_path: Path) -> list[dict[str, str]]:
        if not file_path.exists():
            return []
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            result.append(
                {
                    headers[i]: str(cell) if cell is not None else ""
                    for i, cell in enumerate(row)
                    if i < len(headers)
                }
            )
        return result
